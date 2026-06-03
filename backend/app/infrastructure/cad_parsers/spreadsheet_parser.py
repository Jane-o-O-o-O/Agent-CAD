from io import BytesIO

from app.domain.models.cad_intake import ExtractedContent, ExtractedSourceFile, ExtractedTable
from app.infrastructure.cad_parsers.base import CADFileParser


class SpreadsheetCADParser(CADFileParser):
    name = "spreadsheet"

    async def parse(self, file: ExtractedSourceFile, data: BytesIO) -> ExtractedContent:
        content = ExtractedContent()
        if file.extension != ".xlsx":
            content.uncertain_items.append(
                f"{file.filename} is an .xls file. Configure OfficeConverter to convert .xls to .xlsx before parsing."
            )
            return content

        try:
            from openpyxl import load_workbook
        except ImportError:
            content.uncertain_items.append("openpyxl is not installed; cannot parse .xlsx files.")
            return content

        workbook = load_workbook(data, read_only=True, data_only=True)
        for sheet in workbook.worksheets[:20]:
            rows: list[list[str]] = []
            for row in sheet.iter_rows(max_row=500, max_col=100, values_only=True):
                normalized = ["" if value is None else str(value) for value in row]
                if any(cell.strip() for cell in normalized):
                    rows.append(normalized)
            if rows:
                content.tables.append(
                    ExtractedTable(
                        source_file=file.filename,
                        rows=rows,
                        label=sheet.title,
                        metadata={
                            "sheet": sheet.title,
                            "truncated_rows": sheet.max_row > 500,
                            "truncated_cols": sheet.max_column > 100,
                        },
                    )
                )
        workbook.close()
        return content

